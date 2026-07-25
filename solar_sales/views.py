import json
from decimal import Decimal
from django.db.models import Sum
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from .models import SolarQuotation, SolarQuotationItem, SolarInvoice, SolarProduct, SolarProductCategory, SolarRawMaterialCategory
from .forms import SolarQuotationStep1Form, SolarProductForm
from solar_jobs.models import SolarJob
import openpyxl
from django.http import HttpResponse

# ==========================================
# 📊 Dashboards & Lists
# ==========================================
@login_required
def solar_sales_dashboard(request):
    total_sales = SolarInvoice.objects.filter(status='PAID').aggregate(Sum('grand_total'))['grand_total__sum'] or 0
    pending_balance = SolarInvoice.objects.filter(status='UNPAID').aggregate(Sum('balance_amount'))['balance_amount__sum'] or 0
    draft_qt_count = SolarQuotation.objects.filter(status='DRAFT').count()
    unpaid_inv_count = SolarInvoice.objects.filter(status='UNPAID').count()
    recent_invoices = SolarInvoice.objects.all().order_by('-date', '-id')[:5]

    context = {
        'total_sales': total_sales,
        'pending_balance': pending_balance,
        'draft_qt_count': draft_qt_count,
        'unpaid_inv_count': unpaid_inv_count,
        'recent_invoices': recent_invoices,
    }
    return render(request, 'solar_sales/dashboard.html', context)

@login_required
def solar_quotation_list(request):
    quotations = SolarQuotation.objects.all().order_by('-date', '-id')
    return render(request, 'solar_sales/quotation_list.html', {'quotations': quotations})

# ==========================================
# 📝 ระบบใบเสนอราคา (Knockdown Style)
# ==========================================
@login_required
def solar_quotation_create(request):
    if request.method == 'POST':
        form = SolarQuotationStep1Form(request.POST)
        if form.is_valid():
            qt = form.save(commit=False)
            if hasattr(request.user, 'employee'):
                qt.employee = request.user.employee
            qt.save()
            messages.success(request, f"✅ สร้างเอกสาร {qt.code} เรียบร้อย! กรุณาเพิ่มรายการสินค้า")
            return redirect('solar_quotation_edit', qt_id=qt.id)
    else:
        form = SolarQuotationStep1Form()

    return render(request, 'solar_sales/quotation_form.html', {'form': form})

def calculate_solar_totals(qt):
    subtotal = sum(item.amount for item in qt.items.all())
    qt.subtotal = subtotal
    total_before_vat = subtotal - qt.discount + qt.survey_fee

    if qt.vat_type == 'EXCLUDE':
        qt.vat_amount = total_before_vat * Decimal('0.07')
        qt.grand_total = total_before_vat + qt.vat_amount
    elif qt.vat_type == 'INCLUDE':
        qt.grand_total = total_before_vat
        qt.vat_amount = total_before_vat - (total_before_vat / Decimal('1.07'))
    else: # NONE
        qt.vat_amount = 0
        qt.grand_total = total_before_vat
    qt.save()

@login_required
def solar_quotation_edit(request, qt_id):
    qt = get_object_or_404(SolarQuotation, pk=qt_id)

    if request.method == 'POST':
        if qt.status != 'DRAFT':
            messages.error(request, "❌ ไม่สามารถแก้ไขได้ เนื่องจากเอกสารนี้ถูกอนุมัติหรือล็อกไปแล้ว")
            return redirect('solar_quotation_edit', qt_id=qt.id)

        # 1. จัดการเพิ่มสินค้าหลัก (FG)
        if 'add_item' in request.POST:
            qty = int(request.POST.get('quantity', 1))
            price = Decimal(request.POST.get('price', '0').replace(',', '') or 0)
            product_id = request.POST.get('product_id')
            item_name = request.POST.get('item_name')

            prod = SolarProduct.objects.filter(id=product_id).first() if product_id else None
            SolarQuotationItem.objects.create(
                quotation=qt, product=prod, item_name=item_name or (prod.name if prod else "สินค้าพิเศษ"),
                quantity=qty, unit_price=price, amount=qty * price
            )
            calculate_solar_totals(qt)
            return redirect('solar_quotation_edit', qt_id=qt.id)

        # 2. จัดการเพิ่มของแถม/Upsale (RM)
        elif 'add_upsale' in request.POST:
            qty = int(request.POST.get('upsale_qty', 1))
            price = Decimal(request.POST.get('upsale_price', '0').replace(',', '') or 0)
            desc = request.POST.get('upsale_desc')

            SolarQuotationItem.objects.create(
                quotation=qt, item_name=desc, quantity=qty, unit_price=price, amount=qty * price
            )
            calculate_solar_totals(qt)
            return redirect('solar_quotation_edit', qt_id=qt.id)

        # 3. อัปเดตข้อมูลทั่วไป
        elif 'update_info' in request.POST or 'finish_quote' in request.POST:
            qt.discount = Decimal(request.POST.get('discount', '0').replace(',', '') or 0)
            qt.survey_fee = Decimal(request.POST.get('survey_fee', '0').replace(',', '') or 0)
            qt.payment_terms = request.POST.get('payment_terms', '')
            qt.note = request.POST.get('note', '')
            calculate_solar_totals(qt)
            if 'finish_quote' in request.POST:
                messages.success(request, f"✅ สร้างใบเสนอราคา {qt.code} เสร็จสมบูรณ์แล้ว! (รอผู้อนุมัติ)")
                return redirect('solar_quotation_list')
            return redirect('solar_quotation_edit', qt_id=qt.id)

    # === ส่งข้อมูล JSON ให้ Select2 ===
    products_list = [{'id': p.id, 'name': p.name, 'code': p.code, 'sell_price': float(p.sell_price), 'category_id': p.category_id} for p in SolarProduct.objects.filter(is_active=True, product_type='FG')]
    upsales_list = [{'id': a.id, 'name': a.name, 'code': a.code, 'sell_price': float(a.sell_price), 'category_id': a.rm_category_id} for a in SolarProduct.objects.filter(is_active=True, product_type='RM')]

    return render(request, 'solar_sales/quotation_edit.html', {
        'qt': qt,
        'main_categories': SolarProductCategory.objects.all(),
        'upsale_categories': SolarRawMaterialCategory.objects.all(),
        'products_json': json.dumps(products_list),
        'upsales_json': json.dumps(upsales_list),
        'item_total': sum(i.amount for i in qt.items.all()),
    })

@login_required
def solar_delete_item(request, item_id):
    item = get_object_or_404(SolarQuotationItem, pk=item_id)
    qt = item.quotation
    if qt.status != 'DRAFT':
        messages.error(request, "❌ ไม่สามารถลบรายการได้ เนื่องจากเอกสารนี้ถูกอนุมัติหรือล็อกไปแล้ว")
        return redirect('solar_quotation_edit', qt_id=qt.id)
    item.delete()
    calculate_solar_totals(qt)
    return redirect('solar_quotation_edit', qt_id=qt.id)

@login_required
def solar_quotation_approve(request, qt_id):
    qt = get_object_or_404(SolarQuotation, pk=qt_id)
    qt.status = 'APPROVED'
    qt.save()
    messages.success(request, f"✅ อนุมัติใบเสนอราคา {qt.code} เรียบร้อยแล้ว")
    return redirect('solar_quotation_edit', qt_id=qt.id)

@login_required
def solar_quotation_send_to_center(request, qt_id):
    qt = get_object_or_404(SolarQuotation, pk=qt_id)
    qt.is_deposit_paid = True
    qt.status = 'CONVERTED'
    qt.save()

    first_item = qt.items.first()
    package = first_item.product if first_item else None

    job = SolarJob.objects.create(
        customer=qt.customer,
        salesperson=qt.employee,
        package_sold=package,
        status='DRAFT',
        note=f"📌 สร้างอัตโนมัติจากใบเสนอราคาโซล่า: {qt.code}\nรายละเอียดเพิ่มเติม: {qt.note}"
    )

    if not hasattr(qt, 'solarinvoice'):
        SolarInvoice.objects.create(
            quotation_ref=qt,
            customer=qt.customer,
            grand_total=qt.grand_total,
            balance_amount=qt.grand_total - qt.deposit_amount,
            status='UNPAID' if (qt.grand_total - qt.deposit_amount) > 0 else 'PAID'
        )

    messages.success(request, f"🚀 ส่งงานเข้า Center สำเร็จ! ระบบสร้างรหัสงาน {job.code} และออกใบเสร็จเรียบร้อยแล้ว")
    return redirect('solar_quotation_list')

# ==========================================
# 💰 ระบบการเงิน / Invoice
# ==========================================
@login_required
def solar_invoice_list(request):
    invoices = SolarInvoice.objects.all().order_by('-date', '-id')
    return render(request, 'solar_sales/invoice_list.html', {'invoices': invoices})

@login_required
def solar_invoice_detail(request, inv_id):
    inv = get_object_or_404(SolarInvoice, pk=inv_id)
    if request.method == 'POST':
        inv.balance_amount = 0
        inv.status = 'PAID'
        inv.save()
        messages.success(request, f"✅ ยืนยันการรับชำระเงินบิล {inv.code} ปิดยอดเรียบร้อยแล้ว!")
        return redirect('solar_invoice_detail', inv_id=inv.id)
    return render(request, 'solar_sales/invoice_detail.html', {'inv': inv})

@login_required
def solar_quotation_print(request, qt_id):
    qt = get_object_or_404(SolarQuotation, pk=qt_id)

    # ดึงข้อมูลบริษัทมาโชว์ที่หัวกระดาษ
    from master_data.models import CompanyInfo
    company = CompanyInfo.objects.first()

    # คำนวณราคาสินค้ารวมก่อนหักส่วนลด
    item_total = sum(item.amount for item in qt.items.all())

    return render(request, 'solar_sales/quotation_print.html', {
        'qt': qt,
        'company': company,
        'item_total': item_total
    })

@login_required
def solar_invoice_print(request, inv_id):
    inv = get_object_or_404(SolarInvoice, pk=inv_id)
    return render(request, 'solar_sales/invoice_print.html', {'inv': inv})

# ==========================================
# 📦 คลังสินค้าโซล่า (Inventory & Excel Import)
# ==========================================
@login_required
def solar_inventory_list(request):
    fg_products = SolarProduct.objects.filter(product_type='FG').order_by('-is_active', '-created_at')
    rm_products = SolarProduct.objects.filter(product_type='RM').order_by('-is_active', '-created_at')
    return render(request, 'solar_sales/inventory_list.html', {
        'fg_products': fg_products,
        'rm_products': rm_products
    })

@login_required
def solar_product_create(request):
    default_type = request.GET.get('type', 'FG')
    page_title = 'เพิ่มวัตถุดิบ/อุปกรณ์เสริม' if default_type == 'RM' else 'เพิ่มสินค้า/แพ็กเกจใหม่'

    if request.method == 'POST':
        form = SolarProductForm(request.POST)
        if form.is_valid():
            prod = form.save()
            messages.success(request, f"✅ เพิ่มรายการ '{prod.name}' ลงในคลังสินค้าเรียบร้อยแล้ว")
            return redirect('solar_inventory_list')
        else:
            messages.error(request, "❌ กรุณาตรวจสอบความถูกต้องของข้อมูล")
    else:
        form = SolarProductForm(initial={'product_type': default_type})

    return render(request, 'solar_sales/product_form.html', {'form': form, 'title': page_title})

@login_required
def solar_product_edit(request, pk):
    product = get_object_or_404(SolarProduct, pk=pk)
    if request.method == 'POST':
        form = SolarProductForm(request.POST, instance=product)
        if form.is_valid():
            form.save()
            messages.success(request, f"✅ อัปเดตข้อมูล '{product.name}' เรียบร้อยแล้ว")
            return redirect('solar_inventory_list')
        else:
            messages.error(request, "❌ กรุณาตรวจสอบความถูกต้องของข้อมูล")
    else:
        form = SolarProductForm(instance=product)
    return render(request, 'solar_sales/product_form.html', {'form': form, 'product': product, 'title': f'แก้ไข: {product.name}'})

@login_required
def solar_inventory_download_template(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="inventory_template.xlsx"'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Solar Inventory"
    headers = ['ประเภท (FG หรือ RM)', 'รหัสสินค้า (เว้นว่างได้)', 'หมวดหมู่ (Category)', 'ชื่อสินค้า/แพ็กเกจ', 'หน่วยนับ', 'ราคาทุน', 'ราคาขาย', 'จำนวนคงเหลือ', 'จุดสั่งซื้อ']
    ws.append(headers)
    ws.append(['RM', 'RM-001', 'อุปกรณ์สายไฟ', 'สายไฟ DC 4mm (ตัวอย่าง)', 'เมตร', 15, 25, 100, 20])
    ws.append(['FG', '', 'แพ็กเกจ 5kW', 'แพ็กเกจ 5kW 1 Phase (ตัวอย่าง)', 'ชุด', 100000, 150000, 5, 2])
    wb.save(response)
    return response

@login_required
def solar_inventory_import(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        if not excel_file.name.endswith('.xlsx'):
            messages.error(request, '❌ ไฟล์ไม่ถูกต้อง กรุณาอัปโหลดไฟล์นามสกุล .xlsx เท่านั้น')
            return redirect('solar_inventory_list')

        try:
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            sheet = wb.active
            count_created, count_updated = 0, 0

            for row in sheet.iter_rows(min_row=2, values_only=True):
                product_type = str(row[0]).strip().upper() if row[0] else 'RM'
                code = str(row[1]).strip() if row[1] else None
                cat_name = str(row[2]).strip() if row[2] else None
                name = str(row[3]).strip() if row[3] else None
                unit = str(row[4]).strip() if row[4] else 'ชิ้น'
                cost_price = float(row[5]) if row[5] else 0.0
                sell_price = float(row[6]) if row[6] else 0.0
                stock_qty = float(row[7]) if row[7] else 0.0
                min_level = float(row[8]) if row[8] else 0.0

                if name and name != 'None':
                    if product_type not in ['FG', 'RM']: product_type = 'RM'
                    fg_cat, rm_cat = None, None
                    if cat_name and cat_name != 'None':
                        if product_type == 'FG': fg_cat, _ = SolarProductCategory.objects.get_or_create(name=cat_name)
                        else: rm_cat, _ = SolarRawMaterialCategory.objects.get_or_create(name=cat_name)

                    defaults_data = {
                        'product_type': product_type, 'name': name, 'category': fg_cat, 'rm_category': rm_cat,
                        'unit': unit, 'cost_price': cost_price, 'sell_price': sell_price, 'stock_qty': stock_qty,
                        'min_level': min_level, 'is_active': True
                    }
                    if code and code != 'None':
                        obj, created = SolarProduct.objects.update_or_create(code=code, defaults=defaults_data)
                        if created: count_created += 1
                        else: count_updated += 1
                    else:
                        SolarProduct.objects.create(**defaults_data)
                        count_created += 1

            messages.success(request, f'✅ นำเข้าข้อมูลสำเร็จ! สร้างใหม่ {count_created} รายการ, อัปเดต {count_updated} รายการ')
        except Exception as e:
            messages.error(request, f'❌ เกิดข้อผิดพลาดในการอ่านไฟล์: {str(e)}')
    return redirect('solar_inventory_list')